import openpyxl
import openpyxl
import os
import shutil
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import time
from urllib.parse import urlparse
import requests
from bs4 import BeautifulSoup
import rdflib

"""This script reads all the inputed uris out of a scraped EA excel file (TagsAndNotes),
afterwards, it searches for deadlinks and incorrect anchors.
"""

def get_redirected_url(starting_url):
    """
    Check the final redirected URL of a starting URL.

    Parameters:
    starting_url (str): The URL to start with.

    Returns:
    str: The final redirected URL.
    """
    try:
        response = requests.get(starting_url)
        # Check if the response was successful
        if response.status_code == 200:
            return response.url
        else:
            return f"Error: Request failed with status code {response.status_code}"
    except Exception as e:
        return f"Error: {str(e)}"

# Function to check if a URL is valid
def is_url_valid(url):
    try:
        # Remove trailing slash if present for the initial check
        if url.endswith('/'):
            url = url[:-1]

        response = requests.head(url, allow_redirects=True, timeout=5)
        # If HEAD is not allowed, a GET request is tried
        if response.status_code != 200:
            response = requests.get(url, timeout=5)
        return response.status_code == 200
    except requests.RequestException as e:
        # If the URL with the trailing slash was not valid, check without it
        if url.endswith('/'):
            return is_url_valid(url[:-1])
        print(f"Error with URL {url}: {e}")
        return False


def is_rdf(url):
    return get_redirected_url(url)

def is_ttl(url):
    if '#' in url:
        base_url = base_url.split('#')[0]
        
    return is_url_valid(url)


def check_anchor_in_rdf(url, anchor):
    graph = rdflib.Graph()
    graph.parse(url)
    for s, p, o in graph:
        if anchor == s.split('#')[-1] or anchor == p.split('#')[-1] or anchor == o.split('#')[-1]:
            return True
    return False

def check_anchor_in_html(url, anchor, parser):
    try:
        response = requests.get(url, timeout=5)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, features=parser)
            return bool(soup.find(id=anchor) or soup.find_all(attrs={"name": anchor}))
        return False
    except requests.RequestException as e:
        print(f"Error fetching page for URL {url}: {e}")
        return False

def check_anchor_in_url(url):
    redirected_url = get_redirected_url(url)
    parsed_url = urlparse(redirected_url)
    page_url = parsed_url.scheme + "://" + parsed_url.netloc + parsed_url.path
    anchor = parsed_url.fragment

    if not anchor:
        return True  # No anchor to check

    try:
        response = requests.head(redirected_url, allow_redirects=True, timeout=5)
        content_type = response.headers.get('Content-Type', '')

        if 'rdf' in content_type:
            return check_anchor_in_rdf(page_url, concat('\'', anchor, '\''))
        if 'text/turtle' in content_type or 'ld+json' in content_type:
            return check_anchor_in_rdf(page_url, anchor)
        
        else:
            parser = 'xml' if 'xml' in content_type else 'lxml'
            return check_anchor_in_html(page_url, anchor, parser)

    except requests.RequestException as e:
        print(f"Error fetching page for URL {url}: {e}")
        return False

    
def read_uri_column_from_xlsm(file_path):
    """
    Reads the 'uri' column from the 'TagsAndNotes' sheet of the TagsAndNotes.xlsm file.

    :param file_path: Path to the .xlsm file
    :return: List of URIs from the 'uri' column
    """
    try:
        # Load the workbook and select the specified sheet
        workbook = openpyxl.load_workbook(file_path, keep_vba=True)
        sheet = workbook['TagsAndNotes']

        # Read the 'uri' column (assuming it is the first column)
        uri_column = []
        for row in sheet.iter_rows(min_row=2):  # Assuming the first row is the header
            uri_cell = row[5]  # 'uri' is the sixt column
            uri_column.append(uri_cell.value)

        return uri_column

    except FileNotFoundError:
        print(f"File not found: {file_path}")
        return []
    except KeyError:
        print(f"Sheet 'TagsAndNotes' not found in the workbook.")
        return []
    except Exception as e:
        print(f"An error occurred: {e}")
        return []

def search_for_anchor_in_rdf(rdf_url, anchor):
    try:
        response = requests.get(rdf_url)
        response.raise_for_status()
        if anchor in response.text:
            return True
        else:
            return False

    except requests.RequestException as e:
        print(f"An error occurred while fetching RDF data: {e}")
        return None

def convert_base_url(base_url):
    # Splitting the base URL at '#' to separate the URL part and the anchor
    parts = base_url.split('#')
    if len(parts) != 2:
        raise ValueError("Invalid base URL format. It should contain an anchor.")

    url_part, anchor = parts

    # Construct the RDF URL by inserting 'ontologies/' and appending '.rdf'
    url_parts = url_part.split('/')
    rdf_url = '/'.join(url_parts[:3]) + '/ontologies/' + '/'.join(url_parts[3:]) + '.rdf'

    return rdf_url, anchor

def is_file_open(file_path):
    """
    Check if the file is open by trying to rename it.
    """
    if not os.path.exists(file_path):
        return False  # File does not exist, so it's not open

    try:
        # Attempt to rename the file
        os.rename(file_path, file_path)
        return False
    except:
        return True  # Unable to rename, file is likely open

def change_cell_background_to_red(file_path, sheet_name, row, column):
    """
    Opens an .xlsm file and changes the background color of a specified cell to red.
    """
    try:
        if is_file_open(file_path):
            print("File is currently open. Closing it now.")
            os.system(f'taskkill /f /im excel.exe')

            # Wait a bit to ensure Excel has been closed
            time.sleep(5)

        # Load the workbook
        workbook = openpyxl.load_workbook(file_path, keep_vba=True)
        sheet = workbook[sheet_name]

        # Convert row and column to cell address
        cell_address = f"{get_column_letter(column)}{row}"

        # Change the cell's background color to red
        red_fill = PatternFill(start_color='FFFF0000',
                               end_color='FFFF0000',
                               fill_type='solid')
        cell = sheet[cell_address]
        cell.fill = red_fill

        # Save the changes
        workbook.save(file_path)


    except FileNotFoundError:
        print(f"File not found: {file_path}")
    except KeyError:
        print(f"Sheet '{sheet_name}' not found in the workbook.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Example usage
file_path = 'C:/Users/samue/Documents/OSLO Tools/EA-Excel/TagsAndNotes/TagsAndNotes.xlsm'  # Replace with your file path
sheet_name = 'TagsAndNotes'  # Replace with your sheet name

path = '../TagsAndNotes.xlsm'
uris = read_uri_column_from_xlsm(path)


for i in range(len(uris)):
    if uris[i] is None:
        continue
    if 'def.isotc211.org' in uris[i]:
        rdf_url, anchor = convert_base_url(uris[i])
        if search_for_anchor_in_rdf(rdf_url, anchor):
            continue
        else:
            print('On row :', i, ' Anchor in URI is invalid or not found: ', uris[i])
            change_cell_background_to_red(file_path, sheet_name, i+2, 6)
    else:
        try:
            if is_url_valid(uris[i]):
                    if '#' in uris[i]:
                        if check_anchor_in_url(uris[i]):
                            continue
                        else:
                            print('On row :', i, ' Anchor in URI is invalid or not found: ', uris[i])
                            change_cell_background_to_red(file_path, sheet_name, i+2, 6)
    
            else:
                print('On row :', i, ' we have a dead link ', uris[i])
                change_cell_background_to_red(file_path, sheet_name, i+2, 6)

        except requests.HTTPError as http_err:
            continue
        except Exception as err:
            continue

# Open the file after saving
os.startfile(file_path)