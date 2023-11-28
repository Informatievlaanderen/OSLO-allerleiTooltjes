import openpyxl
import os
import shutil
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import time

def is_file_open(file_path):
    """
    Check if the file is open by trying to rename it.
    """
    if not os.path.exists(file_path):
        return False  # File does not exist, so it's not open

    try:
        # Attempt to rename the file
        os.rename(file_path, file_path)
        return False
    except:
        return True  # Unable to rename, file is likely open

def change_cell_background_to_red(file_path, sheet_name, row, column):
    """
    Opens an .xlsm file and changes the background color of a specified cell to red.
    """
    try:
        if is_file_open(file_path):
            print("File is currently open. Closing it now.")
            os.system(f'taskkill /f /im excel.exe')

            # Wait a bit to ensure Excel has been closed
            time.sleep(5)

        # Load the workbook
        workbook = openpyxl.load_workbook(file_path, keep_vba=True)
        sheet = workbook[sheet_name]

        # Convert row and column to cell address
        cell_address = f"{get_column_letter(column)}{row}"

        # Change the cell's background color to red
        red_fill = PatternFill(start_color='FFFF0000',
                               end_color='FFFF0000',
                               fill_type='solid')
        cell = sheet[cell_address]
        cell.fill = red_fill

        # Save the changes
        workbook.save(file_path)

        # Open the file after saving
        os.startfile(file_path)

    except FileNotFoundError:
        print(f"File not found: {file_path}")
    except KeyError:
        print(f"Sheet '{sheet_name}' not found in the workbook.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Example usage
file_path = 'C:/Users/samue/Documents/OSLO Tools/EA-Excel/TagsAndNotes/TagsAndNotes.xlsm'  # Replace with your file path
sheet_name = 'TagsAndNotes'  # Replace with your sheet name
row = 1  # Replace with the row number you want to change
column = 1  # Replace with the column number you want to change

change_cell_background_to_red(file_path, sheet_name, row, column)


# Open the file after saving
os.startfile(file_path)